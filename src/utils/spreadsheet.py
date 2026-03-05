"""
Reads a CSV or Excel spreadsheet and returns a list of Post objects.

Expected columns (case-insensitive, order doesn't matter):
  Title         – headline / subject
  Content       – body text to post
  Image         – URL(s) or local path(s) to image/video, comma-separated
  Scheduled At  – date-time string, e.g. "2025-06-01 14:30"
  Ready         – YES / NO (only YES rows are processed)
  Platform      – which platform this row targets: mastodon / instagram /
                  facebook / linkedin  (leave blank or omit column to post
                  to all platforms passed on the CLI)
  Posted        – stamped YES automatically after the row is successfully sent
                  (rows with Posted == YES are skipped on future runs)
"""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

# openpyxl is only needed for .xlsx files
try:
    import openpyxl  # noqa: F401
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── Data model ────────────────────────────────────────────────────────────────

@dataclass
class Post:
    title: str
    content: str
    media: list[str] = field(default_factory=list)   # URLs or local paths
    scheduled_at: Optional[datetime] = None
    ready: bool = False
    platform: Optional[str] = None   # e.g. "mastodon", "linkedin", or None = all
    raw: dict = field(default_factory=dict, repr=False)

    @property
    def full_text(self) -> str:
        """Returns title + content combined, ready for posting."""
        if self.title and self.title.lower() not in self.content.lower():
            return f"{self.title}\n\n{self.content}"
        return self.content


# ── Column name normalisation ─────────────────────────────────────────────────

_TITLE_ALIASES    = {"title", "headline", "subject"}
_CONTENT_ALIASES  = {"content", "text", "body", "textual content", "message"}
_MEDIA_ALIASES    = {"image", "images", "media", "video", "image/video", "image/video links"}
_SCHEDULE_ALIASES = {"scheduled at", "date", "time", "date and time", "post date",
                     "publish date", "scheduled_at", "schedule"}
_READY_ALIASES    = {"ready", "publish", "approved", "go", "status"}
_PLATFORM_ALIASES = {"platform", "platforms", "channel", "network", "social media"}
_POSTED_ALIASES   = {"posted", "sent", "done", "published"}


def _normalise(col: str) -> str:
    return col.strip().lower()


def _find_col(headers: list[str], aliases: set[str]) -> Optional[str]:
    for h in headers:
        if _normalise(h) in aliases:
            return h
    return None


# ── Date parsing ──────────────────────────────────────────────────────────────

_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%B %d, %Y %H:%M",
    "%B %d, %Y",
]


def _parse_date(value: str) -> Optional[datetime]:
    if not value:
        return None
    value = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


# ── Row → Post ────────────────────────────────────────────────────────────────

def _row_to_post(row: dict, col_map: dict) -> Optional[Post]:
    def get(aliases_key: str) -> str:
        col = col_map.get(aliases_key)
        return str(row.get(col, "") or "").strip() if col else ""

    ready_val = get("ready").upper()
    if ready_val != "YES":
        return None

    # Skip rows already posted
    if get("posted").upper() == "YES":
        return None

    media_raw = get("media")
    media = [m.strip() for m in media_raw.split(",") if m.strip()] if media_raw else []

    platform_raw = get("platform").lower().strip()

    return Post(
        title=get("title"),
        content=get("content"),
        media=media,
        scheduled_at=_parse_date(get("schedule")),
        ready=True,
        platform=platform_raw if platform_raw else None,
        raw=dict(row),
    )


# ── Public API ────────────────────────────────────────────────────────────────

def load_posts(path: str | Path) -> list[Post]:
    """
    Load posts from a CSV or Excel (.xlsx) file.

    Returns only rows where Ready == YES.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {path}")

    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        rows = _read_excel(path)
    elif ext == ".csv":
        rows = _read_csv(path)
    else:
        raise ValueError(f"Unsupported file format: {ext} (use .csv or .xlsx)")

    if not rows:
        return []

    headers = list(rows[0].keys())
    col_map = {
        "title":    _find_col(headers, _TITLE_ALIASES),
        "content":  _find_col(headers, _CONTENT_ALIASES),
        "media":    _find_col(headers, _MEDIA_ALIASES),
        "schedule": _find_col(headers, _SCHEDULE_ALIASES),
        "ready":    _find_col(headers, _READY_ALIASES),
        "platform": _find_col(headers, _PLATFORM_ALIASES),
        "posted":   _find_col(headers, _POSTED_ALIASES),
    }

    posts = []
    for row in rows:
        post = _row_to_post(row, col_map)
        if post is not None:
            posts.append(post)
    return posts


def _read_csv(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_excel(path: Path) -> list[dict]:
    if not _HAS_OPENPYXL:
        raise ImportError("Install openpyxl to read .xlsx files: pip install openpyxl")
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


# ── Write-back ────────────────────────────────────────────────────────────────

def mark_posted(path: str | Path, post: Post) -> None:
    """
    Stamp Posted = YES on the matching row in an .xlsx file.
    Matches by Title + Content (first 80 chars). Adds a 'Posted' column if absent.
    No-op for CSV files (read-only in this implementation).
    """
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return  # CSV write-back not supported

    if not _HAS_OPENPYXL:
        return

    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    # Find or create the Posted column
    posted_col = None
    for idx, h in enumerate(header_row, start=1):
        if h.lower() in _POSTED_ALIASES:
            posted_col = idx
            break
    if posted_col is None:
        posted_col = len(header_row) + 1
        ws.cell(row=1, column=posted_col, value="Posted")

    # Find title/content columns for matching
    title_col   = next((i+1 for i, h in enumerate(header_row) if h.lower() in _TITLE_ALIASES), None)
    content_col = next((i+1 for i, h in enumerate(header_row) if h.lower() in _CONTENT_ALIASES), None)

    for row in ws.iter_rows(min_row=2):
        row_title   = str(row[title_col-1].value or "").strip()   if title_col   else ""
        row_content = str(row[content_col-1].value or "").strip() if content_col else ""
        if row_title == post.title and row_content[:80] == post.content[:80]:
            ws.cell(row=row[0].row, column=posted_col, value="YES")

    wb.save(path)
