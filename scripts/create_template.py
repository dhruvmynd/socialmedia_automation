"""
Run this script once to generate posts_template.xlsx with the correct columns
and a few example rows.

  python scripts/create_template.py
"""

from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise SystemExit("Install openpyxl first: pip install openpyxl")

HEADERS = ["Title", "Content", "Image", "Scheduled At", "Ready", "Platform", "Posted"]

EXAMPLES = [
    [
        "Lab Update – March (Mastodon)",
        "Exciting news from the lab! We've just published our latest findings on ...",
        "https://example.com/images/lab-photo.jpg",
        "2026-06-01 10:00",
        "YES",
        "mastodon",
        "",
    ],
    [
        "Lab Update – March (LinkedIn)",
        "We are thrilled to share our latest research publication. Our team investigated ...",
        "https://example.com/images/lab-photo.jpg",
        "2026-06-01 10:00",
        "YES",
        "linkedin",
        "",
    ],
    [
        "Conference Reminder",
        "Join us at the annual symposium next week. Register at ...",
        "",
        "2026-06-03 09:00",
        "YES",
        "",
        "",
    ],
    [
        "Draft Post",
        "This post is not ready yet.",
        "",
        "2026-06-10 12:00",
        "NO",
        "",
        "",
    ],
    [
        "Field Trip Photos (Instagram)",
        "Check out these photos from our field trip!",
        "https://example.com/img1.jpg, https://example.com/img2.jpg",
        "",
        "YES",
        "instagram",
        "",
    ],
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
COL_WIDTHS = [30, 60, 50, 22, 10, 14, 10]

out = Path("posts_template.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Posts"

# Headers
for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[cell.column_letter].width = width

# Example rows
for row_idx, row_data in enumerate(EXAMPLES, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save(out)
print(f"Template saved to {out.resolve()}")
print("Columns: Title | Content | Image | Scheduled At | Ready | Platform")
print("Platform: mastodon / instagram / facebook / linkedin  (blank = all)")
print("Set Ready = YES and fill in Scheduled At to queue a post.")
